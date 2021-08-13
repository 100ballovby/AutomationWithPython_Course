import openpyxl as xl

wb = xl.load_workbook('example.xlsx')  # открываю таблицу
sheet = wb['Sheet1']  # сохраняю рабочий лист

fruits_update = {
    'Pears': 20,
    'Apples': 30,
    'Bananas': 20,
}

for row in range(1, sheet.max_row + 1):
    """перебираю все строки таблицы"""
    prod_name = sheet.cell(row, 2).value
    if prod_name in fruits_update:
        sheet.cell(row, 3).value += fruits_update[prod_name]

wb.save('new_example.xlsx')

matrix = {
    'A': ['4/5/15 13:34','4/5/15 3:41','4/6/15 12:46','4/8/15 8:59','4/10/15 2:07','4/10/15 18:10','4/10/15 2:40'],
    'B': ['Apples', 'Cherries','Pears','Oranges','Apples','Bananas','Strawberries']
}
matrix['B'][0] = 'DragonFruit'
print(matrix)
