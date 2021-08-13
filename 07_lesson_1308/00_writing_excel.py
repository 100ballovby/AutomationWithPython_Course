import openpyxl as xl


wb = xl.Workbook()  # создает пустую таблицу excel
print(wb.sheetnames)
active = wb.active  # сохраняю активный рабочий лист

active.title = 'Мое название листа'
print(wb.sheetnames)
wb.create_sheet(title='Второй лист')

wb.create_sheet(index=1, title='Средний лист')
# wb.remove(active)
print(wb.sheetnames)

active['A1'] = 'Мое первое сообщение!'
print(active['A1'].value)

# Сохранение таблицы
wb.save('my_workbook.xlsx')