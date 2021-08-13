from requests import get
from bs4 import BeautifulSoup
import docx
import openpyxl as xl

url = 'https://dominos.by'
res = get(url)

parser = BeautifulSoup(res.text, 'html.parser')

names = []
prices = []

for name in parser.select('.product-card__title'):
    names.append(name.text)

for price in parser.select('.product-card__modification-info-price'):
    prices.append(price.text[:5].replace('.', ','))

pizzas = {}
for i in range(len(prices)):
    pizzas[names[i]] = prices[i]

doc = docx.Document()  # создаю документ Word
wb = xl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Название Пиццы'
sheet['B1'] = 'Стоимость пиццы'

# TODO: доделать нормально цикл записи пицц в таблицу
for key, value in pizzas.items():
    doc.add_paragraph(f'🍕{key}: 💰{value}💰;')
    for row in range(1, len(pizzas) + 1):
        sheet.cell(row, 1).value = f'🍕{key}'
        sheet.cell(row, 2).value = f'🍕{value}'

wb.save('dominos.xlsx')
doc.save('dominos_doc.docx')
